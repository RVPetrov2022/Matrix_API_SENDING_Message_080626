#   Ниже представлен готовый скрипт. Он автоматически определяет, запускается ли логика впервые (создание пространства «ПИК» и комнат с записью в новый файл) или это повторный запуск (валидация токена и сверка ID комнат).
import os
import sys
import openpyxl
import requests

# --- НАСТРОЙКИ ---
MATRIX_URL = "http://192.168.10.15:8008"
USER_TOKEN = "syt_cGV0cm92X3J2MQ_DNeSNAsWtqgzzjBBxCsD_2emR7O"
SPACE_NAME = "ПИК"

INPUT_FILE = "output.xlsx"
OUTPUT_FILE = "output_with_matrix_ids.xlsx"

HEADERS = {
    "Authorization": f"Bearer {USER_TOKEN}",
    "Content-Type": "application/json"
}

def create_matrix_room(name, is_space=False, parent_id=None):
    """Создает комнату или пространство в Matrix"""
    url = f"{MATRIX_URL}/_matrix/client/v3/createRoom"
    
    # Базовые параметры создания
    creation_content = {}
    if is_space:
        creation_content["type"] = "m.space"
        
    payload = {
        "name": name,
        "visibility": "private",
        "preset": "private_chat",
        "creation_content": creation_content
    }
    
    # Отправка запроса на создание
    try:
        response = requests.post(url, headers=HEADERS, json=payload)
        if response.status_code != 200:
            print(f"Ошибка API Matrix ({response.status_code}): {response.text}")
            return None
        room_id = response.json().get("room_id")
        
        # Если это комната внутри Пространства, линкуем её к родителю
        if parent_id and room_id and not is_space:
            link_room_to_space(parent_id, room_id)
            
        return room_id
    except Exception as e:
        print(f"Не удалось подключиться к серверу Matrix: {e}")
        return None

def link_room_to_space(space_id, room_id):
    """Связывает комнату с пространством (Parent-Child отношение)"""
    # 1. Добавляем дочернюю комнату в пространство
    url_child = f"{MATRIX_URL}/_matrix/client/v3/rooms/{space_id}/state/m.space.child/{room_id}"
    payload_child = {"via": [MATRIX_URL.split("//")[-1].split(":")[0]]}
    requests.put(url_child, headers=HEADERS, json=payload_child)

    # 2. Прописываем родителя внутри самой комнаты
    url_parent = f"{MATRIX_URL}/_matrix/client/v3/rooms/{room_id}/state/m.space.parent/{space_id}"
    payload_parent = {"via": [MATRIX_URL.split("//")[-1].split(":")[0]], "canonical": True}
    requests.put(url_parent, headers=HEADERS, json=payload_parent)

def verify_token():
    """Проверяет валидность токена пользователя"""
    url = f"{MATRIX_URL}/_matrix/client/v3/account/whoami"
    try:
        res = requests.get(url, headers=HEADERS)
        if res.status_code == 200 and "petrov_rv1" in res.json().get("user_id", ""):
            return True
        return False
    except:
        return False

def first_run():
    """Первый запуск: создание пространства, комнат и генерация нового файла"""
    print("Режим: Первый запуск. Создание инфраструктуры...")
    
    # 1. Создаем пространство «ПИК»
    print(f"Создаем пространство «{SPACE_NAME}»...")
    space_id = create_matrix_room(SPACE_NAME, is_space=True)
    if not space_id:
        print("Критическая ошибка: Не удалось создать пространство.")
        sys.exit(1)
    print(f"Пространство успешно создано. ID: {space_id}")

    # Загружаем исходный Excel
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active

    # Словарь для отслеживания уже созданных уникальных id_room_xxx
    # Ключ: id_room_xxx, Значение: Matrix Room ID (!xxxx:matrix.org)
    created_rooms = {}

    # Проходим по строкам исходного файла
    for row_idx in range(1, ws.max_row + 1):
        id_room = ws.cell(row=row_idx, column=1).value
        room_name = ws.cell(row=row_idx, column=3).value

        if not id_room or not room_name:
            continue

        # Если эта группа комнат (id_room) еще не создавалась в Matrix
        if id_room not in created_rooms:
            print(f"Создаем комнату для {id_room} ({room_name})...")
            matrix_room_id = create_matrix_room(room_name, is_space=False, parent_id=space_id)
            if matrix_room_id:
                created_rooms[id_room] = matrix_room_id
            else:
                print(f"Ошибка создания комнаты для {id_room}")
                created_rooms[id_room] = "ERROR"

        # Записываем ID комнаты в 5-й столбец текущей строки
        ws.cell(row=row_idx, column=5).value = created_rooms[id_room]

    # Добавляем техническую строку в самый конец файла для проверки токена при повторном запуске
    # Чтобы не портить структуру колонок, запишем метаданные в отдельный скрытый/технический лист или в конец
    meta_sheet = wb.create_sheet(title="Matrix_Metadata")
    meta_sheet["A1"] = "Saved_Token"
    meta_sheet["B1"] = USER_TOKEN
    meta_sheet["A2"] = "Space_ID"
    meta_sheet["B2"] = space_id

    wb.save(OUTPUT_FILE)
    print(f"Все операции завершены. Новый файл сохранен: {OUTPUT_FILE}")

def second_run():
    """Повторный запуск: верификация токена и проверка существующих ID комнат"""
    print("Режим: Повторный запуск. Верификация данных...")
    
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    
    # 1. Проверяем сохраненный токен
    if "Matrix_Metadata" in wb.sheetnames:
        meta_sheet = wb["Matrix_Metadata"]
        saved_token = meta_sheet["B1"].value
        if saved_token != USER_TOKEN:
            print(f"[ВНИМАНИЕ] Текущий токен скрипта НЕ совпадает с токеном из прошлого запуска!")
            print(f"Ожидался: {saved_token}\nПередан:  {USER_TOKEN}")
    else:
        print("[ПРЕДУПРЕЖДЕНИЕ] Метаданные токена в файле отсутствуют.")

    ws = wb.active
    errors_found = False

    # Сверяем ID комнат через Matrix API
    # Чтобы не дергать API для дублирующихся строк, кэшируем проверенные ID
    checked_matrix_ids = {}

    for row_idx in range(1, ws.max_row + 1):
        id_room = ws.cell(row=row_idx, column=1).value
        room_name = ws.cell(row=row_idx, column=3).value
        matrix_room_id = ws.cell(row=row_idx, column=5).value

        if not id_room or not matrix_room_id:
            continue

        if matrix_room_id == "ERROR":
            print(f"[ОШИБКА] Строка {row_idx}: Для {id_room} ранее не удалось создать комнату.")
            errors_found = True
            continue

        if matrix_room_id not in checked_matrix_ids:
            # Проверяем состояние комнаты на сервере Matrix
            url = f"{MATRIX_URL}/_matrix/client/v3/rooms/{matrix_room_id}/state/m.room.name/"
            try:
                res = requests.get(url, headers=HEADERS)
                if res.status_code == 200:
                    server_room_name = res.json().get("name")
                    checked_matrix_ids[matrix_room_id] = {"valid": True, "name": server_room_name}
                else:
                    checked_matrix_ids[matrix_room_id] = {"valid": False, "reason": f"HTTP {res.status_code}"}
            except Exception as e:
                checked_matrix_ids[matrix_room_id] = {"valid": False, "reason": str(e)}

        # Анализируем результат проверки
        status = checked_matrix_ids[matrix_room_id]
        if not status["valid"]:
            print(f"[НЕСОВПАДЕНИЕ] Строка {row_idx} ({id_room}): Комната {matrix_room_id} не найдена на сервере или недоступна. Причина: {status['reason']}")
            errors_found = True
        else:
            if status["name"] != room_name:
                print(f"[ПРЕДУПРЕЖДЕНИЕ] Строка {row_idx} ({id_room}): Название в файле '{room_name}' не совпадает с именем в Matrix '{status['name']}'")
                errors_found = True

    if not errors_found:
        print("Успех: Все проверенные ID комнат совпадают и активны на сервере!")
    else:
        print("Проверка завершена с замечаниями (см. логи выше).")

if __name__ == "__main__":
    # Проверка базовой доступности токена
    if not verify_token():
        print("Критическая ошибка: Токен пользователя невалиден или сервер Matrix недоступен.")
        sys.exit(1)

    # Определяем режим работы по наличию результирующего файла
    if not os.path.exists(OUTPUT_FILE):
        if not os.path.exists(INPUT_FILE):
            print(f"Ошибка: Исходный файл {INPUT_FILE} не найден в текущей директории.")
            sys.exit(1)
        first_run()
    else:
        second_run()
